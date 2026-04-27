"""
excel_exporter.py
Generates a polished, SharePoint-ready Excel file from the events list.
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

COLUMNS = [
    ("Venue",              "venue_name",         28),
    ("Address",            "venue_address",       35),
    ("Event Name",         "event_name",          42),
    ("Type",               "event_type",          13),
    ("Date",               "event_date",          13),
    ("Local Time",         "event_time_local",    12),
    ("GMT+2",              "event_time_gmt2",     10),
    ("Duration",           "duration_scheduled",  11),
    ("Capacity / Tickets", "capacity_or_tickets", 28),
    ("Source URL",         "source_url",          35),
]

TYPE_COLORS = {
    "NFL":           "1A56DB",
    "NBA":           "E3000F",
    "NHL":           "00539C",
    "MLB":           "002D72",
    "MLS":           "19236D",
    "Soccer":        "009900",
    "Concert":       "7C3AED",
    "Entertainment": "D97706",
    "Other":         "6B7280",
}


def _thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def export_to_excel(events: list[dict]) -> io.BytesIO:
    wb = Workbook()

    # ── Events sheet ───────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Events"
    ws.sheet_view.showGridLines = False

    # Header row
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="F8FAFC", name="Calibri", size=11)
    header_border = _thin_border()

    for col_idx, (label, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 26

    # Data rows
    for row_idx, event in enumerate(events, 2):
        even_row = (row_idx % 2 == 0)
        row_fill = PatternFill("solid", fgColor="F1F5F9" if even_row else "FFFFFF")
        border = _thin_border()

        for col_idx, (_, key, _) in enumerate(COLUMNS, 1):
            value = str(event.get(key, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (2, 3, 9, 10)))

            # Colour the Type cell
            if key == "event_type":
                hex_color = TYPE_COLORS.get(value, "6B7280")
                cell.fill = PatternFill("solid", fgColor=hex_color)
                cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif key == "event_date":
                cell.font = Font(bold=True, name="Calibri", size=10)
                cell.fill = row_fill
            elif key in ("event_time_local", "event_time_gmt2"):
                cell.font = Font(name="Courier New", size=10, color="1D4ED8")
                cell.fill = row_fill
            else:
                cell.fill = row_fill

        ws.row_dimensions[row_idx].height = 20

    # Excel Table for easy filtering
    if events:
        last_col = get_column_letter(len(COLUMNS))
        last_row = len(events) + 1
        table = Table(
            displayName="EventsTable",
            ref=f"A1:{last_col}{last_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    ws.freeze_panes = "A2"

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    bold = Font(bold=True, name="Calibri", size=11)
    title_font = Font(bold=True, name="Calibri", size=14, color="0F172A")

    ws2["A1"] = "Event Radar – Summary Report"
    ws2["A1"].font = title_font
    ws2.row_dimensions[1].height = 28

    ws2["A3"] = "Generated"
    ws2["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A3"].font = bold

    ws2["A4"] = "Total Events"
    ws2["B4"] = len(events)
    ws2["A4"].font = bold

    # Counts by type
    from collections import Counter
    counts = Counter(e.get("event_type", "Other") for e in events)
    ws2["A6"] = "Breakdown by Type"
    ws2["A6"].font = bold

    for i, (etype, cnt) in enumerate(sorted(counts.items()), 7):
        ws2.cell(row=i, column=1, value=etype)
        ws2.cell(row=i, column=2, value=cnt)
        hex_c = TYPE_COLORS.get(etype, "6B7280")
        ws2.cell(row=i, column=1).fill = PatternFill("solid", fgColor=hex_c)
        ws2.cell(row=i, column=1).font = Font(color="FFFFFF", name="Calibri", size=10, bold=True)

    # ── Write to buffer ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
